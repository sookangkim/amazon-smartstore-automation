#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
아마존 크롤링 데이터를 네이버 스마트스토어 실제 업로드 가능한 형식으로 변환

실제 스마트스토어 일괄등록 Excel 템플릿에 맞는 형식으로 변환

작성일: 2025년 8월 1일
버전: v2.0 - 실제 업로드 형식
"""

import pandas as pd
import json
import os
import re
from datetime import datetime
from typing import Dict, List, Optional
import logging
import unicodedata

# 번역 모듈 import
try:
    from translator import ProductTranslator
    TRANSLATION_AVAILABLE = True
except ImportError:
    TRANSLATION_AVAILABLE = False
    print("번역 모듈을 찾을 수 없습니다. 영어 원본으로 진행됩니다.")

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class SmartstoreUploader:
    """네이버 스마트스토어 실제 업로드 형식 변환기"""
    
    def __init__(self, enable_translation=True):
        """변환기 초기화"""
        self.usd_to_krw = 1350  # 환율
        self.markup_percentage = 50  # 기본 마진율 50%
        
        # 번역기 초기화
        self.enable_translation = enable_translation and TRANSLATION_AVAILABLE
        if self.enable_translation:
            try:
                self.translator = ProductTranslator()
                logger.info("번역 기능이 활성화되었습니다.")
            except Exception as e:
                logger.warning(f"번역기 초기화 실패: {e}")
                self.enable_translation = False
        else:
            self.translator = None
        
        # 네이버 스마트스토어 실제 카테고리 코드 (2025년 기준)
        self.category_codes = {
            # 뷰티/화장품 카테고리
            'serum': '50000169',             # 스킨케어 > 에센스/세럼
            'cream': '50000167',             # 스킨케어 > 크림
            'skincare': '50000166',          # 스킨케어 
            'beauty': '50000166',            # 스킨케어
            'cosmetics': '50000166',         # 스킨케어
            'hyaluronic': '50000169',        # 히알루론산 = 에센스/세럼
            'retinol': '50000167',           # 레티놀 = 크림
            'vitamin_c': '50000169',         # 비타민C = 에센스/세럼
            
            # 건강기능식품
            'protein': '50006674',           # 건강기능식품
            'vitamin': '50006674',           # 건강기능식품 
            'supplement': '50006674',        # 건강기능식품
            
            # 기타 카테고리
            'baby': '50002436',              # 베이비/출산용품
            'pet': '50002439',               # 반려동물용품
            'home': '50000131',              # 생활용품
            'kitchen': '50000156',           # 주방용품
            'tech': '50000128',              # 전자제품
            'fashion': '50000001',           # 패션의류
            'office': '50000145'             # 문구/오피스
        }
        
        # 뷰티/스킨케어를 기본값으로 변경 (현재 상품들이 뷰티 제품)
        self.default_category_code = '50000169'  # 에센스/세럼
        
        # 네이버 스마트스토어 세부 카테고리 매핑 (최종 카테고리까지)
        self.detailed_category_mapping = {
            '50000169': {  # 에센스/세럼
                'category_path': '뷰티 > 스킨케어 > 에센스/세럼',
                'category_id': '50000169',
                'leaf_category_id': '50000169'
            },
            '50000167': {  # 크림
                'category_path': '뷰티 > 스킨케어 > 크림',
                'category_id': '50000167', 
                'leaf_category_id': '50000167'
            },
            '50006674': {  # 건강기능식품
                'category_path': '건강 > 건강기능식품 > 비타민/미네랄',
                'category_id': '50006674',
                'leaf_category_id': '50006674'
            }
        }
    
    def clean_text_for_excel(self, text: str) -> str:
        """Excel 파일용 텍스트 정리 (이모지 및 특수문자 제거)"""
        if not text:
            return ""
        
        # 유니코드 이모지 제거
        text = re.sub(r'[\U00010000-\U0010ffff]', '', text)
        
        # 다양한 이모지 패턴 제거
        emoji_pattern = re.compile("["
                                 u"\U0001F600-\U0001F64F"  # emoticons
                                 u"\U0001F300-\U0001F5FF"  # symbols & pictographs
                                 u"\U0001F680-\U0001F6FF"  # transport & map symbols
                                 u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
                                 u"\U00002700-\U000027BF"  # Dingbats
                                 u"\U0000FE00-\U0000FE0F"  # Variation Selectors
                                 u"\U00002600-\U000026FF"  # Miscellaneous Symbols
                                 u"\U00002B00-\U00002BFF"  # Miscellaneous Symbols and Arrows
                                 "]+", flags=re.UNICODE)
        text = emoji_pattern.sub('', text)
        
        # 기타 문제가 될 수 있는 문자들 제거
        text = re.sub(r'[™®©℠]', '', text)
        text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', text)  # 제어 문자 제거
        
        # 다중 공백 정리
        text = re.sub(r'\s+', ' ', text).strip()
        
        return text
    
    def clean_and_translate_title(self, title: str) -> Dict[str, str]:
        """상품명 정리 및 번역"""
        # 특수문자 정리
        cleaned_title = re.sub(r'[™®©]', '', title)
        cleaned_title = re.sub(r'\s+', ' ', cleaned_title).strip()
        
        # 브랜드명 추출
        brand = ""
        product_name = cleaned_title
        
        words = cleaned_title.split()
        if len(words) > 1:
            potential_brand = words[0]
            if len(potential_brand) > 2 and potential_brand[0].isupper():
                brand = potential_brand
                product_name = ' '.join(words[1:])
        
        # 번역 적용
        if self.enable_translation and self.translator:
            try:
                korean_product_name = self.translator.translate_product_title(product_name)
                if brand:
                    final_title = f"{brand} {korean_product_name}"
                else:
                    final_title = korean_product_name
                
                # 스마트스토어 제목 길이 제한 (100자)
                if len(final_title) > 100:
                    final_title = final_title[:97] + "..."
                
                return {
                    'brand': brand,
                    'product_name': korean_product_name,
                    'final_title': final_title,
                    'original_title': cleaned_title
                }
            except Exception as e:
                logger.warning(f"제목 번역 실패: {e}")
        
        # 번역 실패 시 원본 사용
        if len(product_name) > 100:
            product_name = product_name[:97] + "..."
        
        return {
            'brand': brand,
            'product_name': product_name,
            'final_title': cleaned_title,
            'original_title': cleaned_title
        }
    
    def calculate_korean_price(self, usd_price: str, margin_rate: int = None) -> int:
        """검증된 한국 판매가격 계산 (쿠팡 변환기 로직 적용)"""
        try:
            price_usd = float(str(usd_price).replace('$', '').replace(',', ''))
            if price_usd <= 0:
                return 0
            
            # 환율 적용
            krw_price = price_usd * self.usd_to_krw
            
            # 쿠팡 변환기의 검증된 가격 조정 로직 적용
            # 1.6배 적용 후 800원 단위 반올림
            adjusted_price = krw_price * 1.6
            
            # 800원 단위로 반올림 (쿠팡 변환기와 동일)
            if adjusted_price % 100 != 0:
                adjusted_price = round(adjusted_price / 800) * 800
            
            # 최소 가격 보장 (1000원)
            final_price = max(adjusted_price, 1000)
            
            return int(final_price)
            
        except (ValueError, TypeError) as e:
            logger.error(f"가격 계산 오류: {e}")
            return 0
    
    def get_category_code(self, product_title: str, category: str = '') -> str:
        """상품명과 카테고리를 기반으로 올바른 카테고리 코드 반환"""
        title_lower = product_title.lower() if product_title else ""
        category_lower = category.lower() if category else ""
        
        # 상품명 기반 키워드 매칭 (우선순위)
        if any(word in title_lower for word in ['serum', '세럼', 'essence', '에센스']):
            return '50000169'  # 에센스/세럼
        elif any(word in title_lower for word in ['cream', '크림', 'moisturizer', '모이스처']):
            return '50000167'  # 크림
        elif any(word in title_lower for word in ['retinol', '레티놀']):
            return '50000167'  # 레티놀 크림
        elif any(word in title_lower for word in ['hyaluronic', '히알루론산']):
            return '50000169'  # 히알루론산 세럼
        elif any(word in title_lower for word in ['vitamin c', '비타민c', 'vitamin-c']):
            return '50000169'  # 비타민C 세럼
        
        # 카테고리 기반 키워드 매칭 (보조)
        for keyword, code in self.category_codes.items():
            if keyword in title_lower or keyword in category_lower:
                return code
        
        return self.default_category_code
    
    def convert_to_smartstore_upload_format(self, amazon_data: List[Dict]) -> pd.DataFrame:
        """아마존 데이터를 스마트스토어 실제 업로드 형식으로 변환 (검증된 89개 필드)"""
        converted_products = []
        
        # 입력 데이터 검증
        if not amazon_data:
            logger.error("변환할 아마존 데이터가 없습니다.")
            return pd.DataFrame()
        
        logger.info(f"변환 시작: {len(amazon_data)}개 상품")
        
        for i, product in enumerate(amazon_data, 1):
            try:
                # 기본 데이터 검증
                if not isinstance(product, dict):
                    logger.warning(f"상품 {i}: 올바르지 않은 데이터 형식, 건너뜀")
                    continue
                
                # 필수 필드 확인
                required_fields = ['title', 'price_usd']
                missing_fields = [field for field in required_fields if not product.get(field)]
                if missing_fields:
                    logger.warning(f"상품 {i}: 필수 필드 누락 ({missing_fields}), 건너뜀")
                    continue
                title_info = self.clean_and_translate_title(product.get('title', ''))
                sale_price = self.calculate_korean_price(product.get('price_usd', 0))
                category_code = self.get_category_code(title_info['final_title'], product.get('category', ''))
                
                # 상품 설명 번역 및 보완
                description = ""
                if self.enable_translation and self.translator:
                    try:
                        original_desc = product.get('description', '') or product.get('features', '')
                        if original_desc:
                            if isinstance(original_desc, list):
                                translated_features = self.translator.translate_product_features(original_desc[:3])
                                description = " / ".join(translated_features)
                            else:
                                description = self.translator.translate_product_description(original_desc)
                        
                        # 설명이 없거나 짧을 경우 기본 설명 추가
                        if not description or len(description.strip()) < 50:
                            # 상품명 기반 기본 설명 생성 (이모지 제거)
                            product_name = title_info['final_title']
                            if 'serum' in product_name.lower() or '세럼' in product_name:
                                description = f"{product_name}\\n\\n* 프리미엄 스킨케어 세럼\\n* 피부에 깊은 영양과 수분 공급\\n* 건강하고 윤기있는 피부로 가꾸어 드립니다\\n\\n* 안전한 해외직구 상품\\n* 빠른 배송 서비스 제공"
                            elif 'cream' in product_name.lower() or '크림' in product_name:
                                description = f"{product_name}\\n\\n* 프리미엄 스킨케어 크림\\n* 피부에 깊은 보습과 영양 공급\\n* 부드럽고 촉촉한 피부로 가꾸어 드립니다\\n\\n* 안전한 해외직구 상품\\n* 빠른 배송 서비스 제공"
                            else:
                                description = f"{product_name}\\n\\n* 프리미엄 뷰티 제품\\n* 피부 건강을 위한 전문 케어\\n* 아름답고 건강한 피부로 가꾸어 드립니다\\n\\n* 안전한 해외직구 상품\\n* 빠른 배송 서비스 제공"
                        
                        # 길이 제한 (32700자 - Excel 제한)
                        if len(description) > 32700:
                            description = description[:32697] + "..."
                    except Exception as e:
                        logger.warning(f"설명 번역 실패: {e}")
                        # 번역 실패 시에도 기본 설명 제공 (이모지 제거)
                        product_name = title_info['final_title']
                        description = f"{product_name}\\n\\n* 프리미엄 뷰티 제품\\n* 피부 건강을 위한 전문 케어\\n* 아름답고 건강한 피부로 가꾸어 드립니다\\n\\n* 안전한 해외직구 상품\\n* 빠른 배송 서비스 제공"
                
                # 모든 텍스트 필드 정리
                final_title = self.clean_text_for_excel(title_info['final_title'])
                brand = self.clean_text_for_excel(title_info['brand'])
                description = self.clean_text_for_excel(description)
                
                # 카테고리 상세 정보 가져오기
                category_detail = self.detailed_category_mapping.get(category_code, {
                    'category_path': '뷰티 > 스킨케어 > 에센스/세럼',
                    'category_id': '50000169',
                    'leaf_category_id': '50000169'
                })
                
                # 검증된 네이버 스마트스토어 완전 필수 필드 구조 (단일 행 헤더 적용)
                smartstore_product = {
                    # 핵심 상품 정보 (0-6)
                    '판매자상품코드': f'AMZ_{i:04d}',
                    '카테고리코드': category_code,
                    '상품명': final_title,
                    '상품상태': '신상품',
                    '판매가': sale_price,
                    '부가세': '과세상품',
                    '재고수량': 999,
                    
                    # 필수 추가 필드들
                    '최종카테고리선택': category_detail['leaf_category_id'],
                    '구매평노출여부': 'Y',
                    '상품문의노출여부': 'Y', 
                    '리뷰작성가능여부': 'Y',
                    '판매상태': '판매중',
                    '전시상태': '전시',
                    '성인인증': 'N',
                    '청소년이용불가': 'N',
                    
                    # 옵션 관련 필드 (7-16) - 옵션 없는 단순 상품으로 설정
                    '옵션형태': '단순상품',
                    '옵션명': '',
                    '옵션값': '',
                    '옵션가': '',
                    '옵션재고수량': '',
                    '직접입력옵션': '',
                    '추가상품명': '',
                    '추가상품값': '',
                    '추가상품가': '',
                    '추가상품재고수량': '',
                    
                    # 이미지 및 설명 (17-19) - 유효한 이미지 URL 설정
                    '대표이미지': '',
                    '추가이미지': '',
                    '상세설명': description,
                    
                    # 제조 및 원산지 정보 (20-28) - 정확한 원산지코드 사용
                    '브랜드': brand,
                    '제조사': brand or '해외제조사',
                    '제조일자': '2024-01-01',
                    '유효일자': '2030-12-31',
                    '원산지코드': 'US',
                    '수입사': '',
                    '복수원산지여부': 'N',
                    '원산지직접입력': '미국',
                    '미성년자구매여부': 'N',
                    
                    # 배송비 템플릿코드 추가
                    '배송비템플릿코드': '1',  # 기본 배송비 템플릿
                    
                    # 배송 관련 정보 (29-45) - 정확한 배송 설정
                    '배송방법': '택배',
                    '기본배송비': 3000,
                    '배송비유형': '유료',
                    '배송비결제방식': '선결제',
                    '출고지': '서울',
                    '배송업체': 'CJ대한통운',
                    '배송기간': '1~3일',
                    '조건부무료상품판매가합계': '',
                    '수량별부과수량': '',
                    '구간별2구간수량': '',
                    '구간별3구간수량': '',
                    '구간별3구간배송비': '',
                    '구간별추가배송비': '',
                    '반품배송비': '',
                    '교환배송비': '',
                    '지역별차등배송비': '',
                    '별도설치비': '',
                    
                    # 상품정보제공고시 (46-50) - 상품정보제공고시 템플릿코드 사용  
                    '상품정보제공고시템플릿코드': '50000169',  # 에센스/세럼 템플릿
                    '상품정보제공고시품명': final_title,
                    '상품정보제공고시모델명': final_title[:30] + '_' + f'AMZ_{i:04d}',
                    '상품정보제공고시인증허가사항': 'FDA 승인 시설에서 제조',
                    '상품정보제공고시제조자': brand or '해외제조사',
                    '상품정보제공고시제조국': '미국',
                    '상품정보제공고시사용기한': '제품 표기 참조',
                    '상품정보제공고시사용법': '제품 설명서 참조',
                    '상품정보제공고시주의사항': '사용 전 패치테스트 권장',
                    
                    # A/S 관련 (51-54) - A/S 템플릿코드 사용
                    'AS템플릿코드': '1',  # 기본 A/S 템플릿
                    'AS담당자명': '고객센터',
                    'AS전화번호': '010-2291-4080',
                    'AS안내': 'A/S 관련 문의는 판매자에게 연락바랍니다. 해외 직구 상품으로 A/S는 제한적입니다.',
                    '판매자특이사항': '해외 직구 상품입니다',
                    
                    # 할인 및 포인트 관련 (55-68) - 모든 필드 비워두기 (사용 안함)
                    '즉시할인값기본할인': '',
                    '즉시할인단위기본할인': '',
                    '모바일즉시할인값': '',
                    '모바일즉시할인단위': '',
                    '복수구매할인조건값': '',
                    '복수구매할인조건단위': '',
                    '복수구매할인값': '',
                    '복수구매할인단위': '',
                    '상품구매시포인트지급값': '',
                    '상품구매시포인트지급단위': '',
                    '텍스트리뷰작성시지급포인트': '',
                    '포토동영상리뷰작성시지급포인트': '',
                    '한달사용텍스트리뷰작성시지급포인트': '',
                    '한달사용포토동영상리뷰작성시지급포인트': '',
                    
                    # 기타 인증 및 상품정보 (69-88) - 화장품 관련 필수 정보
                    '가전효율등급': '',
                    '효율등급인증기관': '',
                    '케어라벨인증유형': '',
                    '상품정보제공고시색상': '제품 참조',
                    '상품정보제공고시소재': '화장품',
                    '상품정보제공고시사이즈': '제품 상세 참조',
                    '상품정보제공고시동백사이즈': '',
                    '상품정보제공고시동백노출': '',
                    '상품정보제공고시수리방법': '',
                    '사이즈상품군': '일반',
                    '사이즈사이즈명': 'FREE',
                    '사이즈상세사이즈': '제품 상세 참조',
                    '사이즈모델명': f'MODEL{i:04d}',
                    
                    # 참고용 데이터 (업로드에는 포함되지 않음)
                    '상품설명_참고': description,
                    '아마존평점': product.get('rating', 0),
                    '아마존리뷰수': product.get('review_count', 0),
                    '아마존USD가격': product.get('price_usd', 0),
                    '아마존원본제목': title_info['original_title'],
                    '이미지URL': product.get('image_url', ''),
                    '브랜드_참고': title_info['brand'],
                    '수집일시': product.get('crawl_timestamp', datetime.now().isoformat())
                }
                
                converted_products.append(smartstore_product)
                logger.info(f"상품 {i} 변환 완료: {title_info['final_title'][:30]}...")
                
            except Exception as e:
                logger.error(f"상품 {i} 변환 실패: {e}")
                # 에러 세부사항 로깅 (디버깅용)
                logger.debug(f"상품 데이터: {product}")
                continue
        
        # 결과 검증
        if not converted_products:
            logger.error("변환된 상품이 없습니다. 모든 상품에서 오류가 발생했습니다.")
            return pd.DataFrame()
        
        logger.info(f"변환 완료: {len(converted_products)}개 상품 성공")
        
        # DataFrame 생성 시 에러 방지
        try:
            df = pd.DataFrame(converted_products)
            
            # DataFrame 기본 검증
            if df.empty:
                logger.error("DataFrame이 비어있습니다.")
                return pd.DataFrame()
            
            # 중요 컬럼 존재 확인 (단일 행 헤더 적용)
            required_columns = ['판매자상품코드', '카테고리코드', '상품명', '판매가']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                logger.error(f"필수 컬럼 누락: {missing_columns}")
                # 누락된 컬럼 추가 (기본값으로)
                for col in missing_columns:
                    df[col] = ''
            
            return df
            
        except Exception as e:
            logger.error(f"DataFrame 생성 실패: {e}")
            return pd.DataFrame()
    
    def create_upload_file(self, df: pd.DataFrame, output_path: str = None) -> str:
        """스마트스토어 업로드용 Excel 파일 생성 (단일 시트)"""
        # 입력 검증
        if df is None or df.empty:
            logger.error("생성할 데이터가 없습니다. DataFrame이 비어있습니다.")
            return None
        
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"smartstore_upload_{timestamp}.xlsx"
        
        logger.info(f"Excel 파일 생성 시작: {output_path}")
        logger.info(f"데이터 크기: {df.shape[0]}행 {df.shape[1]}열")
        
        try:
            # 스마트스토어 업로드용 단일 행 헤더 컬럼 목록 (줄바꿈 문제 해결)
            upload_columns = [
                '판매자상품코드', '카테고리코드', '상품명', '상품상태', '판매가', '부가세', 
                '재고수량', '최종카테고리선택', '구매평노출여부', '상품문의노출여부', '리뷰작성가능여부',
                '판매상태', '전시상태', '성인인증', '청소년이용불가',
                '옵션형태', '옵션명', '옵션값', '옵션가', '옵션재고수량', 
                '직접입력옵션', '추가상품명', '추가상품값', '추가상품가', '추가상품재고수량',
                '대표이미지', '추가이미지', '상세설명', '브랜드', '제조사', '제조일자', 
                '유효일자', '원산지코드', '수입사', '복수원산지여부', '원산지직접입력', '미성년자구매여부',
                '배송비템플릿코드', '배송방법', '기본배송비', '배송비유형', '배송비결제방식', '출고지', '배송업체', '배송기간',
                '조건부무료상품판매가합계', '수량별부과수량', '구간별2구간수량', '구간별3구간수량',
                '구간별3구간배송비', '구간별추가배송비', '반품배송비', '교환배송비',
                '지역별차등배송비', '별도설치비',
                '상품정보제공고시템플릿코드', '상품정보제공고시품명', '상품정보제공고시모델명', 
                '상품정보제공고시인증허가사항', '상품정보제공고시제조자', 
                '상품정보제공고시제조국', '상품정보제공고시사용기한', 
                '상품정보제공고시사용법', '상품정보제공고시주의사항',
                'AS템플릿코드', 'AS담당자명', 'AS전화번호', 'AS안내', '판매자특이사항',
                '즉시할인값기본할인', '즉시할인단위기본할인', '모바일즉시할인값',
                '모바일즉시할인단위', '복수구매할인조건값', '복수구매할인조건단위',
                '복수구매할인값', '복수구매할인단위', '상품구매시포인트지급값',
                '상품구매시포인트지급단위', '텍스트리뷰작성시지급포인트',
                '포토동영상리뷰작성시지급포인트', '한달사용텍스트리뷰작성시지급포인트',
                '한달사용포토동영상리뷰작성시지급포인트', '가전효율등급',
                '효율등급인증기관', '케어라벨인증유형', '상품정보제공고시색상',
                '상품정보제공고시소재', '상품정보제공고시사이즈', '상품정보제공고시동백사이즈',
                '상품정보제공고시동백노출', '상품정보제공고시수리방법',
                '사이즈상품군', '사이즈사이즈명', '사이즈상세사이즈', '사이즈모델명'
            ]
            
            # 업로드용 데이터프레임 생성 (검증된 컬럼만)
            upload_df = df[upload_columns].copy()
            
            # openpyxl로 Excel 파일 생성 (인코딩 문제 해결)
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = '일괄등록'
            
            # 헤더 스타일 설정 (네이버 스마트스토어 표준 형식)
            header_font = Font(name='맑은 고딕', size=10, bold=True)
            header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 단일 행 헤더 작성 (다중 행 헤더 문제 해결)
            for col_idx, header in enumerate(upload_columns, 1):
                # 헤더에서 줄바꿈 문자 제거하여 단일 행으로 만들기
                clean_header = header.replace('\r\n', '').replace('\n', '')
                cell = ws.cell(row=1, column=col_idx, value=clean_header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # 데이터 작성
            for row_idx, (_, row) in enumerate(upload_df.iterrows(), 2):
                for col_idx, value in enumerate(row, 1):
                    # 텍스트 정리 적용
                    if isinstance(value, str):
                        value = self.clean_text_for_excel(value)
                    
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 컬럼 너비 자동 조정
            for col_idx, header in enumerate(upload_columns, 1):
                max_len = max(len(str(header)), 15)  # 최소 15자
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)
            
            # 파일 저장
            wb.save(output_path)
            
            # 참고용 정보는 별도 파일로 생성
            reference_path = output_path.replace('.xlsx', '_참고용.xlsx')
            self._create_reference_file(df, reference_path)
            
            logger.info(f"스마트스토어 업로드 파일 생성 완료: {output_path}")
            logger.info(f"참고용 정보 파일 생성 완료: {reference_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"파일 생성 실패: {e}")
            return None
    
    def _create_reference_file(self, df: pd.DataFrame, reference_path: str):
        """참고용 정보 파일 생성"""
        try:
            # 참고용 컬럼들 (업로드용과 분리)
            reference_columns = [
                '카테고리코드', '상품명', '판매가', '재고수량', 'A/S 전화번호',
                '상품설명_참고', '아마존평점', '아마존리뷰수', '아마존USD가격', 
                '아마존원본제목', '이미지URL', '브랜드_참고', '수집일시'
            ]
            
            # 참고용 데이터프레임 생성
            reference_df = df[reference_columns].copy()
            
            # 참고용 파일 생성
            with pd.ExcelWriter(reference_path, engine='openpyxl') as writer:
                reference_df.to_excel(writer, index=False, sheet_name='전체정보')
                
                # 시트 서식 설정
                worksheet = writer.sheets['전체정보']
                
                # 컬럼 너비 자동 조정
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.warning(f"참고용 파일 생성 실패: {e}")
    
    def convert_file(self, input_file: str, output_file: str = None, margin_rate: int = None) -> str:
        """파일 변환 메인 함수"""
        logger.info(f"스마트스토어 업로드 형식 변환 시작: {input_file}")
        
        # 마진율 설정
        if margin_rate:
            self.markup_percentage = margin_rate
        
        # 아마존 데이터 로드
        try:
            if input_file.endswith('.json'):
                with open(input_file, 'r', encoding='utf-8') as f:
                    amazon_data = json.load(f)
            elif input_file.endswith('.csv'):
                df = pd.read_csv(input_file, encoding='utf-8-sig')
                amazon_data = df.to_dict('records')
            else:
                raise ValueError("지원되지 않는 파일 형식입니다. JSON 또는 CSV 파일을 사용해주세요.")
        except Exception as e:
            logger.error(f"파일 로드 실패: {e}")
            return None
        
        if not amazon_data:
            raise ValueError("아마존 데이터를 로드할 수 없습니다.")
        
        logger.info(f"로드된 상품 수: {len(amazon_data)}개")
        
        # 스마트스토어 업로드 형식으로 변환
        upload_df = self.convert_to_smartstore_upload_format(amazon_data)
        
        if upload_df.empty:
            logger.error("변환된 상품이 없습니다. 모든 상품 변환에 실패했습니다.")
            return None
        
        logger.info(f"변환된 상품 수: {len(upload_df)}개")
        
        # 업로드 파일 생성
        output_path = self.create_upload_file(upload_df, output_file)
        
        if not output_path:
            logger.error("Excel 파일 생성에 실패했습니다.")
            return None
        
        logger.info(f"스마트스토어 업로드 파일 생성 완료: {output_path}")
        return output_path

def main():
    """테스트 실행 함수"""
    uploader = SmartstoreUploader()
    
    # 최신 아마존 크롤링 파일 찾기
    import glob
    amazon_files = glob.glob("amazon_products_*.json")
    
    if not amazon_files:
        print("아마존 크롤링 파일을 찾을 수 없습니다.")
        return
    
    # 가장 최신 파일 선택
    latest_file = max(amazon_files, key=os.path.getctime)
    print(f"변환할 파일: {latest_file}")
    
    try:
        output_file = uploader.convert_file(latest_file)
        print(f"✅ 스마트스토어 업로드 파일 생성 완료: {output_file}")
        print("📁 '스마트스토어_업로드용' 시트를 스마트스토어에 업로드하세요!")
        print("📋 '전체정보_참고용' 시트에서 추가 정보를 확인할 수 있습니다.")
    except Exception as e:
        print(f"❌ 변환 실패: {e}")

if __name__ == "__main__":
    main()